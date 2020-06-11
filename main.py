from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# resize image (By Tomvon: https://stackoverflow.com/questions/273946/how-do-i-resize-an-image-using-pil-and-maintain-its-aspect-ratio)
basewidth = 256
img = Image.open('image.jpg')

# what percentage basewidth is of the original width (img.size[0])
wpercent = (basewidth / float(img.size[0]))
hsize = int((float(img.size[1]) * float(wpercent)))
img = img.resize((basewidth, hsize), Image.ANTIALIAS)
img.save('resized_image.jpg')

img = Image.open('resized_image.jpg')
img_width = img.size[0]
img_height = img.size[1]

pixel_data = img.load()

wb = Workbook()
ws = wb.active

# map rgb values of each individual pixel as background of respective cell
print("Plotting the values...")
for w in range(1, img_width):
    for h in range(1, img_height):
        # convert rgb to hex
        color = ('%02x%02x%02x' % pixel_data[w-1, h-1]).upper()

        # make the cells square (uses standard units; not pixels.)
        ws.row_dimensions[h].height = 9.75
        ws.column_dimensions[get_column_letter(w)].width = 1.75

        # fill color
        ws.cell(row=h, column=w, value="").fill = PatternFill(
            start_color=color, end_color=color, fill_type="solid")

print("Saving Spreadsheet...")
wb.save('test.xlsx')
